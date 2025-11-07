# log/views.py
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.utils.timezone import make_aware
from .models import LogEntry
from django.shortcuts import render

@login_required
def page(request):
    action_choices = [{"value": a, "label": b} for a, b in LogEntry.Action.choices]
    return render(
        request,
        "log.html",
        {
            "username": getattr(request.user, "first_name", "") or request.user.get_username(),
            "action_choices": action_choices,
        },
    )

# helper date-range
def _parse_range_str(s: str):
    if not s:
        return (None, None)
    parts = [p.strip() for p in s.split("to")]
    if len(parts) != 2:
        return (None, None)
    try:
        start = datetime.strptime(parts[0], "%Y-%m-%d")
        end = datetime.strptime(parts[1], "%Y-%m-%d")
        return (make_aware(start), make_aware(end))
    except Exception:
        return (None, None)

def _filter_logs(request):
    qs = LogEntry.objects.select_related("user").all()

    user_q = (request.GET.get("user") or "").strip()
    if user_q:
        qs = qs.filter(
            Q(user__username__icontains=user_q) |
            Q(user__first_name__icontains=user_q) |
            Q(user__last_name__icontains=user_q)
        )

    action = (request.GET.get("action") or "").strip()
    if action:
        qs = qs.filter(action=action)

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(details__icontains=q)

    dr = (request.GET.get("daterange") or "").strip()
    start, end = _parse_range_str(dr)
    if start and end:
        qs = qs.filter(created_at__date__range=(start.date(), end.date()))

    return qs.order_by("-created_at")

@login_required
def api_entries(request):
    qs = _filter_logs(request)
    limit = int(request.GET.get("limit", 100))
    offset = int(request.GET.get("offset", 0))
    total = qs.count()
    items = [{
        "user": (le.user.get_full_name() or le.user.get_username()),
        "action": le.get_action_display() if hasattr(le, "get_action_display") else le.action,
        "details": le.details,
        "date": le.created_at.strftime("%Y-%m-%d %H:%M"),
    } for le in qs[offset:offset+limit]]
    return JsonResponse({"total": total, "items": items})


try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    _XLSX_OK = True
except Exception:
    _XLSX_OK = False

@login_required
def api_download(request):
    if not _XLSX_OK:
        return JsonResponse({"error": "openpyxl not installed"}, status=500)

    qs = _filter_logs(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Log"

    # header style
    header_fill = PatternFill(start_color="102B86", end_color="102B86", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["User", "Action", "Details", "Date"]
    for idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_alignment

    r = 2
    for le in qs:
        ws.cell(row=r, column=1, value=(le.user.get_full_name() or le.user.get_username()))
        ws.cell(row=r, column=2, value=(le.get_action_display() if hasattr(le, "get_action_display") else le.action))
        ws.cell(row=r, column=3, value=le.details)
        ws.cell(row=r, column=4, value=le.created_at.strftime("%Y-%m-%d %H:%M"))
        r += 1

    # auto width
    for col in ws.columns:
        max_len = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_len + 2, 60)

    from io import BytesIO
    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)

    resp = HttpResponse(
        buff.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="activity_log.xlsx"'
    return resp
